#!/usr/bin/python3
# -*- coding: utf-8 -*-

"""
fetches mortality data from Destatis
see https://www.destatis.de/DE/Themen/Querschnitt/Corona/Gesellschaft/bevoelkerung-sterbefaelle.html
data: https://www.destatis.de/DE/Themen/Gesellschaft-Umwelt/Bevoelkerung/Sterbefaelle-Lebenserwartung/Tabellen/sonderauswertung-sterbefaelle.html;jsessionid=3B59CB1FA0C08C059243535606A41FBF.internet8721
"""

__author__ = "Dr. Torben Menke"
__email__ = "https://entorb.net"
__license__ = "GPL"

# Built-in/Generic Imports

# external packages
import pandas as pd
import openpyxl

import datetime as dt

# import csv
import urllib.request

from pandas.core.frame import DataFrame
import numpy as np

# my helper modules
import helper


# 1. read my covid data

# read my data
df0 = pd.read_csv("data/de-states/de-state-DE-total.tsv", sep="\t")

# extract only date and Death_New columns
df2 = pd.DataFrame()
df2["Date"] = df0["Date"]
df2["Deaths_Covid"] = df0["Deaths_New"]
del df0

# drop deaths of last 4 weeks, as they are not final yet
df2["DateAsDate"] = pd.to_datetime(df2["Date"], format="%Y-%m-%d")
date_4w = dt.date.today() - dt.timedelta(weeks=3)
# print(date_4w)
df2.loc[df2["DateAsDate"].dt.date >= date_4w, "Deaths_Covid"] = None
# print(df2.tail(30))


# remove 29.2.
df2 = df2[~df2["Date"].isin(("2020-02-29", "2024-02-29", "2028-02-29"))]

# ensure first row is from 02.01. (for prepending only 1 missing day)
assert df2.iloc[0]["Date"] == "2020-01-02", (
    "Error of start date, expecting 2020-01-02, got : " + df2.iloc[0]["Date"]
)

# prepend 1.1.2020
l = [0] * 1  # 1 day
df1 = pd.DataFrame(data={"Deaths_Covid": l})

df3 = DataFrame()
df3["Deaths_Covid"] = df1["Deaths_Covid"].append(df2["Deaths_Covid"], ignore_index=True)
# df4 = pd.concat([df1, df3]).reset_index(drop=True)
del df1, df2

df3["Deaths_Covid_roll"] = (
    df3["Deaths_Covid"].rolling(window=7, min_periods=1).mean().round(1)
)
# rolling takes NAN values into account, so I need to overwrite them
df3["Deaths_Covid_roll"] = np.where(
    df3["Deaths_Covid"].isnull(), np.nan, df3["Deaths_Covid_roll"]
)
# print(df3.tail(30))

df_covid_2020 = (
    df3[0 : 1 * 365]
    .reset_index(drop=True)
    .rename(
        columns={
            "Deaths_Covid": "Deaths_Covid_2020",
            "Deaths_Covid_roll": "Deaths_Covid_2020_roll",
        },
        errors="raise",
    )
)

df_covid_2021 = (
    df3[1 * 365 : 2 * 365]
    .reset_index(drop=True)
    .rename(
        columns={
            "Deaths_Covid": "Deaths_Covid_2021",
            "Deaths_Covid_roll": "Deaths_Covid_2021_roll",
        },
        errors="raise",
    )
)


# 2. fetch and parse Excel of mortality data from Destatis

excelFile = "cache/de-mortality.xlsx"


if not helper.check_cache_file_available_and_recent(
    fname=excelFile,
    max_age=1800,
    verbose=False,  # as file is stored in cache folder which is not part of the commit, we can use the caching here
):
    url = "https://www.destatis.de/DE/Themen/Gesellschaft-Umwelt/Bevoelkerung/Sterbefaelle-Lebenserwartung/Tabellen/sonderauswertung-sterbefaelle.xlsx?__blob=publicationFile"
    filedata = urllib.request.urlopen(url)
    datatowrite = filedata.read()
    with open(excelFile, mode="wb") as f:
        f.write(datatowrite)


# data_only : read values instead of formulas
workbookIn = openpyxl.load_workbook(excelFile, data_only=True)
sheetIn = workbookIn["D_2016_2022_Tage"]

l_dates = []
l_deaths2022 = []
l_deaths2021 = []
l_deaths2020 = []
l_deaths2019 = []
l_deaths2018 = []
l_deaths2017 = []
l_deaths2016 = []
for col in range(2, 368):
    day = sheetIn.cell(column=col, row=9).value
    # we skip the 29.02. for each year
    if day == "29.02.":
        continue
    l_dates.append(day)
    l_deaths2022.append(sheetIn.cell(column=col, row=10).value)
    l_deaths2021.append(sheetIn.cell(column=col, row=11).value)
    l_deaths2020.append(sheetIn.cell(column=col, row=12).value)
    l_deaths2019.append(sheetIn.cell(column=col, row=13).value)
    l_deaths2018.append(sheetIn.cell(column=col, row=14).value)
    l_deaths2017.append(sheetIn.cell(column=col, row=15).value)
    l_deaths2016.append(sheetIn.cell(column=col, row=16).value)


data = zip(
    l_dates,
    l_deaths2016,
    l_deaths2017,
    l_deaths2018,
    l_deaths2019,
    l_deaths2020,
    l_deaths2021,
    l_deaths2022,
)

df = pd.DataFrame(
    data, columns=["Day", "2016", "2017", "2018", "2019", "2020", "2021", "2022"]
)

# df["2022"] = df["2022"].fillna(0)
# df["2022"].astype(int)
# df["2022"] = df["2022"].replace(0, np.nan)

# problem: rolling is extrapolating by 6 days into the future
# , closed= did not fix it
df["2016_roll"] = df["2016"].rolling(window=7, min_periods=1).mean().round(1)
df["2017_roll"] = df["2017"].rolling(window=7, min_periods=1).mean().round(1)
df["2018_roll"] = df["2018"].rolling(window=7, min_periods=1).mean().round(1)
df["2019_roll"] = df["2019"].rolling(window=7, min_periods=1).mean().round(1)
df["2020_roll"] = df["2020"].rolling(window=7, min_periods=1).mean().round(1)
df["2021_roll"] = df["2021"].rolling(window=7, min_periods=1).mean().round(1)
df["2022_roll"] = df["2022"].rolling(window=7, min_periods=1).mean().round(1)
df["2016_2019_mean"] = df.iloc[:, [1, 2, 3, 4]].mean(axis=1)  # not column 0 = day
df["2016_2019_mean_roll"] = (
    df["2016_2019_mean"].rolling(window=7, min_periods=1).mean().round(1)
)

df["2016_2019_roll_max"] = df.iloc[:, [8, 9, 10, 11]].max(axis=1)
df["2016_2019_roll_min"] = df.iloc[:, [8, 9, 10, 11]].min(axis=1)

df = df.join(df_covid_2020).join(df_covid_2021)

df.to_csv("data/de-mortality.tsv", sep="\t", index=False, line_terminator="\n")


# print(df[["2022", "2022_roll"]].head(17))
